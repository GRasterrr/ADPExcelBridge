import tkinter as tk
import webbrowser
from tkinter import messagebox
import requests
from openpyxl import load_workbook
import os
from tkinter import filedialog
from tkinter import ttk
from tkinter import font
from tkinterdnd2 import DND_FILES, TkinterDnD
import json

class DictionaryCreator:
    def __init__(self, root):
        self.root = root
        self.root.title("Albion data project - excel bridge")
        self.root.geometry("700x500")
        self.root.option_add("*Background", "#262626")
        self.root.option_add("*Foreground", "#d6d6d6")

        # Фиксированные параметры для всех словарей
        self.FIXED_PARAMETERS = [
            "item",  # Название (обязательное)
            "region",  # Описание
            "city",  # Категория
            "quality",  # Приоритет (низкий, средний, высокий)
            "stat",  # Статус (активный, завершенный, отложенный)
            "excel_cell",  # Ячейка Excel для экспорта
            "excel_sheet"  # Лист Excel (по умолчанию первый лист)
        ]

        # Список для хранения словарей
        self.dictionaries_list = []

        # Текущий файл Excel
        self.current_excel_file = None

        self.last_selection = None

        # Создаем основной фрейм
        main_frame = tk.Frame(root, padx=10, pady=5)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Фрейм для управления Excel файлом
        excel_frame = tk.Frame(main_frame, relief=tk.GROOVE, padx=10, pady=5, bd=1)
        excel_frame.pack(fill=tk.X)

        tk.Label(excel_frame, text="Excel file:", font=("Arial", 10, "bold")).pack(anchor="w")

        # Информация о текущем файле
        self.file_info_label = tk.Label(excel_frame, text="File not chosen", font=("Arial", 9), foreground="red")
        self.file_info_label.pack(anchor="w")

        # Кнопки управления Excel
        file_button_frame = tk.Frame(excel_frame)
        file_button_frame.pack(fill=tk.X, pady=5)

        def save_items():
            with open('data.json', 'w') as fp:
                json.dump(self.dictionaries_list, fp)

        def drop(event):
            if ".json" in event.data:
                with open(event.data, 'r') as fp:
                    data = json.load(fp)
                    self.dictionaries_list = data
                    self.update_display()
            else:
                self.select_excel_file(file_path=event.data)

        root.drop_target_register(DND_FILES)
        root.dnd_bind("<<Drop>>", drop)

        tk.Button(file_button_frame,
                  text="Bridging Excel file",
                  command=self.select_excel_file,
                  #bg="#bdbdbd"
                  ).pack(side=tk.LEFT, padx=5)

        # Фрейм для ввода имени словаря
        item_frame = tk.Frame(main_frame)
        item_frame.pack(fill=tk.X, pady=5)

        def openhyperlink(url):
            webbrowser.open_new(url)

        # Область для отображения списка словарей
        tk.Label(main_frame, text="Bridged items:", font=("Arial", 11, "bold")).pack(anchor="w", pady=(5, 5))

        # Фрейм для списка словарей с прокруткой
        list_frame = tk.Frame(main_frame)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        self.list_frame_right = tk.Frame(list_frame)
        self.list_frame_right.pack(side=tk.RIGHT)

        # Полоса прокрутки
        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Listbox для отображения словарей
        self.dict_listbox = tk.Listbox(
            list_frame,
            yscrollcommand=scrollbar.set,
            font=("Arial", 10),
            selectbackground="lightblue",
            selectmode=tk.SINGLE,
            height=8,
            activestyle="none",
            borderwidth=0,
            highlightbackground="#666666",
            highlightcolor="#666666",
            exportselection=False
        )
        self.dict_listbox.pack(side=tk.LEFT,fill=tk.BOTH, expand=True)

        scrollbar.config(command=self.dict_listbox.yview)

        # Фрейм для кнопок управления
        button_frame = tk.Frame(main_frame)
        button_frame.pack(anchor="w", fill=tk.X)

        save_items_btn = tk.Button(button_frame, text="Save items", command=save_items)
        save_items_btn.pack(side="left", anchor="w")

        button_frame_right = tk.Frame(button_frame)
        button_frame_right.pack(side="right")

        self.create_button = tk.Button(
            self.list_frame_right,
            text="Create item",
            command=self.create_dictionary,
            # bg="lightblue",
            font=("Arial", 12),
            width=15
        )
        self.create_button.pack(pady=5, side=tk.BOTTOM)

        self.delete_button = tk.Button(
            button_frame_right,
            text="Delete",
            command=self.delete_selected_dict,
            # bg="lightcoral",
            state=tk.DISABLED
        )
        self.delete_button.pack(padx=5, side="left")

        # Кнопка экспорта в Excel
        self.export_button = tk.Button(
            button_frame_right,
            text="Export to Excel",
            command=self.export_to_excel,
            # bg="lightseagreen",
            fg="white",
            state=tk.DISABLED
        )
        self.export_button.pack(padx=5, side="left")

        # Привязываем выбор элемента к активации кнопок
        self.dict_listbox.bind("<<ListboxSelect>>", self.on_selection_change)

        list_frame_params = tk.Frame(self.list_frame_right)
        list_frame_params.pack(side=tk.TOP)

        def update(*args):
            selection = self.dict_listbox.curselection()
            if not selection:
                return
            index = selection[0]
            if index >= len(self.dictionaries_list):
                return
            new_dict = {
                "item": self.item_entry.get(),  # Название (обязательное)
                "region": self.region_combobox.get(),  # Описание
                "city": self.city_combobox.get(),  # Категория
                "quality": self.quality_combobox.get(),  # Приоритет (низкий, средний, высокий)
                "stat": self.stat_combobox.get(),  # Статус (активный, завершенный, отложенный)
                "excel_cell": self.excelcell_entry.get(),  # Ячейка Excel для экспорта
                "excel_sheet": self.excelsheet_combobox.get(),
            }
            self.dictionaries_list[index] = new_dict
            self.update_display()


        item_frame = tk.Frame(list_frame_params)
        item_frame.pack(pady=5)
        item_label = tk.Label(item_frame, text="Item" + ":", anchor="w", width=10, foreground="#80a3d1", cursor="hand2")
        item_label.pack(side=tk.LEFT)
        underlineFont = font.Font(item_label, item_label.cget("font"))
        underlineFont.configure(underline=True)
        item_label.configure(font=underlineFont)
        item_label.bind("<Button-1>", lambda e: openhyperlink("https://github.com/ao-data/ao-bin-dumps/blob/master/formatted/items.txt"))
        self.item_var = tk.StringVar(value="")
        self.item_var.trace_add("write", update)
        self.item_entry = ttk.Entry(item_frame, width=23, textvariable=self.item_var)
        self.item_entry.pack(fill=tk.X, expand=True, side=tk.LEFT)

        region_frame = tk.Frame(list_frame_params)
        region_frame.pack(pady=5)
        tk.Label(region_frame, text="Region" + ":", anchor="w", width=10).pack(side=tk.LEFT)
        self.region_var = tk.StringVar(value="")
        self.region_var.trace_add("write", update)
        self.region_combobox = ttk.Combobox(region_frame, values=["europe", "east", "west"], textvariable=self.region_var)
        self.region_combobox.pack(fill=tk.X, expand=True)

        city_frame = tk.Frame(list_frame_params)
        city_frame.pack(pady=5)
        tk.Label(city_frame, text="City" + ":", anchor="w", width=10).pack(side=tk.LEFT)
        self.city_var = tk.StringVar(value="")
        self.city_var.trace_add("write", update)
        self.city_combobox = ttk.Combobox(city_frame, textvariable=self.city_var, values=["Thetford", "Fort Sterling", "Lymhurst", "Bridgewatch", "Martlock", "	Caerleon", "Brecilien"])
        self.city_combobox.pack(fill=tk.X, expand=True)

        quality_frame = tk.Frame(list_frame_params)
        quality_frame.pack(pady=5)
        tk.Label(quality_frame, text="Quality" + ":", anchor="w", width=10).pack(side=tk.LEFT)
        self.quality_var = tk.StringVar(value="")
        self.quality_var.trace_add("write", update)
        self.quality_combobox = ttk.Combobox(quality_frame, textvariable=self.quality_var, values=["0", "1", "2", "3", "4"])
        self.quality_combobox.pack(fill=tk.X, expand=True)

        stat_frame = tk.Frame(list_frame_params)
        stat_frame.pack(pady=5)
        tk.Label(stat_frame, text="Stat" + ":", anchor="w", width=10).pack(side=tk.LEFT)
        self.stat_var = tk.StringVar(value="")
        self.stat_var.trace_add("write", update)
        self.stat_combobox = ttk.Combobox(stat_frame, textvariable=self.stat_var, values=["item_id", "city", "quality", "sell_price_min", "sell_price_min_date", "sell_price_max", "sell_price_max_date", "buy_price_min", "buy_price_min_date", "buy_price_max", "buy_price_max_date"])
        self.stat_combobox.pack(fill=tk.X, expand=True)

        excelcell_frame = tk.Frame(list_frame_params)
        excelcell_frame.pack(pady=5)
        tk.Label(excelcell_frame, text="ExcelCell" + ":", anchor="w", width=10).pack(side=tk.LEFT)
        self.excelcell_var = tk.StringVar(value="")
        self.excelcell_var.trace_add("write", update)
        self.excelcell_entry = ttk.Entry(excelcell_frame, width=23, textvariable=self.excelcell_var)
        self.excelcell_entry.pack(fill=tk.X, expand=True)

        excelsheet_frame = tk.Frame(list_frame_params)
        excelsheet_frame.pack(pady=5)
        tk.Label(excelsheet_frame, text="ExcelSheet" + ":", anchor="w", width=10).pack(side=tk.LEFT)
        self.excelsheet_var = tk.StringVar(value="")
        self.excelsheet_var.trace_add("write", update)
        self.excelsheet_combobox = ttk.Combobox(excelsheet_frame, textvariable=self.excelsheet_var)
        self.excelsheet_combobox.pack(fill=tk.X, expand=True)


    def select_excel_file(self, file_path=None):
        """Выбор существующего Excel файла"""
        if file_path is None:
            file_path = filedialog.askopenfilename(
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Choose Excel file"
            )

        if file_path:
            try:
                # Проверяем, что файл существует и доступен
                if os.path.exists(file_path):
                    # Пробуем загрузить файл для проверки
                    wb = load_workbook(file_path)
                    wb.close()

                    self.current_excel_file = file_path
                    self.file_info_label.config(
                        text=f"Selected file: {os.path.basename(file_path)}",
                        foreground="green"
                    )
                    self.export_button.config(state=tk.NORMAL)
                else:
                    messagebox.showerror("Ошибка", "File doesnt exist")
                sheet_names = []
                if self.current_excel_file:
                    try:
                        wb = load_workbook(self.current_excel_file)
                        sheet_names = wb.sheetnames
                        wb.close()
                    except:
                        pass
                self.excelsheet_combobox['values'] = sheet_names

            except Exception as e:
                messagebox.showerror("Ошибка", f"Cant open a file:\n{str(e)}")

    def edit_params(self, dictionary, index):
        self.item_var.set(dictionary["item"])
        self.region_var.set(dictionary["region"])
        self.city_var.set(dictionary["city"])
        self.quality_var.set(dictionary["quality"])
        self.stat_var.set(dictionary["stat"])
        self.excelcell_var.set(dictionary["excel_cell"])
        self.excelsheet_var.set(dictionary["excel_sheet"])

    def on_selection_change(self, event):
        """Активирует/деактивирует кнопки при выборе элемента"""
        selection = self.dict_listbox.curselection()
        if not selection:
            return
        index = selection[0]
        if index >= len(self.dictionaries_list):
            return
        dictionary = self.dictionaries_list[index]
        self.last_selection = index
        self.edit_params(dictionary, index)

        if self.dict_listbox.curselection():
            self.delete_button.config(state=tk.NORMAL)
        else:
            self.delete_button.config(state=tk.NORMAL)

    def delete_selected_dict(self):
        """Удаляет выбранный словарь"""
        selection = self.dict_listbox.curselection()
        if not selection:
            return

        index = selection[0]
        if index >= len(self.dictionaries_list):
            return

        del self.dictionaries_list[index]
        self.update_display()

    def create_dictionary(self):
        """Создает новый словарь с фиксированными параметрами"""
        item = self.item_entry.get().strip()
        region = self.region_combobox.get().strip()
        city = self.city_combobox.get().strip()
        quality = self.quality_combobox.get().strip()
        stat = self.stat_combobox.get().strip()
        excelcell = self.excelcell_entry.get().strip()
        excelsheet = self.excelsheet_combobox.get().strip()

        # Создаем словарь с фиксированными параметрами
        new_dict = {param: "" for param in self.FIXED_PARAMETERS}
        new_dict["item"] = item
        new_dict["region"] = region
        new_dict["city"] = city
        new_dict["quality"] = quality
        new_dict["stat"] = stat
        new_dict["excel_cell"] = excelcell
        new_dict["excel_sheet"] = excelsheet

        # Добавляем в список
        self.dictionaries_list.append(new_dict)
        self.update_display()

    def export_to_excel(self):
        """Экспортирует данные словарей в существующий Excel файл"""
        if not self.dictionaries_list:
            messagebox.showwarning("Warning", "Item list is empty")
            return

        if not self.current_excel_file:
            messagebox.showwarning("Warning", "Choose or create excel file")
            return

        try:
            # Загружаем существующий файл
            wb = load_workbook(self.current_excel_file)

            exported_count = 0
            errors = []

            for i, dictionary in enumerate(self.dictionaries_list):
                excel_cell = dictionary.get('excel_cell', '').strip().upper()
                excel_sheet = dictionary.get('excel_sheet', 'Sheet1').strip()

                if not excel_cell:
                    errors.append(f"Item '{dictionary.get('item', 'Untitled')}': no cell")
                    continue

                # Проверяем корректность формата ячейки
                if not self.is_valid_excel_cell(excel_cell):
                    errors.append(
                        f"Item '{dictionary.get('item', 'Untitled')}': incorrect cell '{excel_cell}'")
                    continue

                # Получаем или создаем лист
                if excel_sheet in wb.sheetnames:
                    ws = wb[excel_sheet]
                else:
                    ws = wb.create_sheet(excel_sheet)
                    errors.append(f"New list: '{excel_sheet}'")

                # Данные для экспорта
                response = requests.get(f'https://{dictionary.get('region')}.albion-online-data.com/api/v2/stats/prices/{dictionary.get('item')}?locations={dictionary.get('city')}&qualities={str(int(dictionary.get('quality'))+1)}')
                export_data = f"{response.json()[0][f'{dictionary.get('stat')}']}"
                if export_data.isdigit() and export_data.isascii():
                    export_data = int(export_data)

                # Записываем данные в указанную ячейку
                ws[excel_cell] = export_data
                exported_count += 1

            # Сохраняем изменения в файле
            wb.save(self.current_excel_file)
            wb.close()

            # Формируем сообщение о результате
            message = f"Exported!\n: {os.path.basename(self.current_excel_file)}\n"
            message += f"Exported count: {exported_count}/{len(self.dictionaries_list)}"

            if errors:
                message += "\n\nErrors,warnings:\n" + "\n".join(errors[:3])  # Показываем первые 3 ошибки
                if len(errors) > 3:
                    message += f"\n... and more {len(errors) - 3} errors"

            messagebox.showinfo("Export result", message)

        except Exception as e:
            messagebox.showerror("Error", f"Unable to export:\n{str(e)}")

    def is_valid_excel_cell(self, cell_reference):
        """Проверяет корректность формата ячейки Excel"""
        import re
        pattern = r'^[A-Z]{1,3}[1-9]\d{0,6}$'
        return bool(re.match(pattern, cell_reference))

    def update_display(self):
        """Обновляет отображение списка словарей"""
        yscrolllevel = self.dict_listbox.yview()
        self.dict_listbox.delete(0, tk.END)

        for dictionary in self.dictionaries_list:
            item = dictionary.get('item', 'Untitled')
            excel_cell = dictionary.get('excel_cell', 'Not given')
            excel_sheet = dictionary.get('excel_sheet', '')

            display_text = f"{item} Cell: {excel_sheet} {excel_cell}"
            self.dict_listbox.insert(tk.END, display_text)
        self.dict_listbox.selection_set(self.last_selection)
        self.dict_listbox.yview_moveto(yscrolllevel[0])

        # Деактивируем кнопки если нет выбора
        if not self.dict_listbox.curselection():
            self.delete_button.config(state=tk.DISABLED)

# Создаем и запускаем приложение
if __name__ == "__main__":
    root = TkinterDnD.Tk()
    app = DictionaryCreator(root)
    root.mainloop()